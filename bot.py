import time
import imaplib
import email
import os
import openpyxl
from openpyxl import Workbook, load_workbook
# from email.header import decode_header
from bs4 import BeautifulSoup
from seleniumwire import webdriver
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service 
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
import logging
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from dotenv import load_dotenv


log_file_path = './log/logs.log'

# Check if the log file exists
if not os.path.exists(log_file_path):
    with open(log_file_path, 'w') as file:
        file.write('')

# Configure the logging module
logging.basicConfig(
    filename=log_file_path,  # Specify the log file name and path
    level=logging.ERROR,            # Set the logging level (DEBUG, INFO, WARNING, ERROR, CRITICAL)
    format='%(asctime)s [%(levelname)s] - %(message)s',
)


# Load env data
load_dotenv()

username = os.getenv('USERNAME')
password = os.getenv('PASSWORD')
hostname = os.getenv('HOSTNAME')
port = os.getenv('PORT')

# Create new options for every driver instance
def get_options():
    options=uc.ChromeOptions()

    options.add_experimental_option("detach", True)
    options.add_argument("--temp-profile")
    options.add_argument("--headless")
    options.add_argument(f'--proxy-server=https://{username}:{password}@{hostname}:{port}')
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-extensions")
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--start-maximized')
    options.add_experimental_option("prefs", {
        "profile.default_content_setting_values.media_stream_mic": 1,
        "profile.default_content_setting_values.media_stream_camera": 1,
        "profile.default_content_setting_values.geolocation": 0,
        "profile.default_content_setting_values.notifications": 1
    })
    return options

def getDriver():
    # Create a new instance of the Chrome driver
    try:
        service = Service(executable_path='./chromedriver')
        driver = webdriver.Chrome(service=service,options=get_options())
    except:
        driver = webdriver.Chrome(options=get_options())

    driver.implicitly_wait(20)
    actions = ActionChains(driver)

    return driver, actions


class IGBot:
    def __init__(self, username, password, newUsername, newEmail, newEmailPassword) -> None:
        self.username = username
        self.password = password
        self.newEmail = newEmail
        self.newUsername = newUsername
        self.newEmailPassword = newEmailPassword

        # Keep track of the session attempts
        self.attempts = 1

    def loginIG(self, driver):
        
        # Instagram page
        driver.get('https://instagram.com/accounts/login/')

        # Wait for the username input field to be present and input username
        username_input = WebDriverWait(driver,  10).until(
            EC.presence_of_element_located((By.NAME, 'username'))
        )
        username_input.send_keys(self.username)

        # Locate the password input field and enter the password
        password_input = driver.find_element(By.NAME, 'password')
        password_input.send_keys(self.password)

        # Submit the login form
        password_input.send_keys(Keys.RETURN)
        

    def changeEmail(self) -> str:
        # Open a new window
        driver, actions = getDriver()
        self.loginIG(driver) #pass the n

        try:
            # Check if the error messages displayed
            incorrect_password = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Sorry, your password was incorrect. Please double-check your password.')]"))
            )  

            nameSignal = "Instagram username or password incorrect"
            driver.quit()
            del driver

            return nameSignal
        except TimeoutException as e:          
            # Load the link for settings
            driver.get("https://accountscenter.instagram.com/personal_info/contact_points/?contact_point_type=email&dialog_type=add_contact_point")

        # Input a new email address
        newEmailForm = WebDriverWait(driver, 50).until(
            EC.presence_of_element_located((By.TAG_NAME, 'input'))
        )
        newEmailForm.send_keys(self.newEmail)

        # Click on the checkbox to accept the terms, tab to the done button
        checkButton = driver.find_element(By.NAME, 'noform')
        checkButton.click()
        checkButton.send_keys(Keys.TAB)

        # Click on the Done button to submit information
        time.sleep(3)
        actions.send_keys(Keys.RETURN)
        actions.perform()

        #Get the OTP from the mail and input in the field
        try:
            otp = getOTP(self.newEmail, self.newEmailPassword)
        except Exception as e:
            logging.error(f"An error occurred: {str(e)}", exc_info=True)
            emailSignal = ("Failed to get OTP.")
            driver.quit()
            del driver

            return emailSignal

        # Select field
        try:
            label = WebDriverWait(driver, 50).until(
                EC.presence_of_element_located((By.XPATH, "//*[text()='Enter confirmation code']"))
            )
            otpID = label.get_attribute('for')
        except:
            emailSignal = ("OTP field not found.")
            driver.quit()
            del driver

            return emailSignal

        # Enter otp and click next
        driver.find_element(By.ID, otpID).send_keys(otp)
        actions.send_keys(Keys.TAB)
        actions.send_keys(Keys.TAB)
        actions.send_keys(Keys.TAB)
        time.sleep(2)
        actions.send_keys(Keys.ENTER)

        time.sleep(3)

        # Refresh to ensure the name propagates around instagram
        refresh_count = 3

        for _ in range(refresh_count):
            # Refresh the page
            driver.refresh()

            # Optional: Add a delay between each refresh
            time.sleep(2)
            emailSignal = "Email Change Done"

        driver.close()
        driver.quit()

        return emailSignal


    def changeName(self) -> str:  
        driver, actions = getDriver()
             
        # Login the email
        self.loginIG(driver) 

        try:
            # Check if the error messages displayed
            incorrect_password = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Sorry, your password was incorrect. Please double-check your password.')]"))
            )  

            nameSignal = "Instagram username or password incorrect"
            driver.quit()
            del driver

            return nameSignal
        
        except TimeoutException as e:          
            # Load the link for settings
            logging.error(f"An error occurred: {str(e)}", exc_info=True)
            driver.get("https://accountscenter.instagram.com/profiles")
    
        #Get the user link and load the page
        try:
            links = WebDriverWait(driver, 50).until(
                EC.presence_of_all_elements_located((By.XPATH, "//*[@role='link']"))
            )
            time.sleep(15)
            userLink = links[10].get_attribute("href")
            driver.get(f"{userLink}username/manage/")

        except Exception as e:
            logging.error(f"An error occurred: {str(e)}", exc_info=True)
            nameSignal = ("Could not get user link")
            driver.quit()
            del driver

            return nameSignal

        # Click on the username settings and replace it with the new username
        newUsernameInput = WebDriverWait(driver, 50).until(
            EC.presence_of_element_located((By.TAG_NAME, 'input'))
        )
        newUsernameInput.send_keys(Keys.CONTROL + "a")
        newUsernameInput.send_keys(Keys.DELETE)
        newUsernameInput.send_keys(self.newUsername)

        # Moves the cursor to the Done button and clicks 
        time.sleep(3)
        newUsernameInput.click()
        actions.send_keys(Keys.TAB)
        actions.send_keys(Keys.TAB)
        time.sleep(3)
        actions.send_keys(Keys.ENTER)
        actions.perform()

        time.sleep(8)

        # Refresh to ensure the name propagates around instagram
        refresh_count = 3

        for _ in range(refresh_count):
            # Refresh the page
            driver.refresh()

            # Optional: Add a delay between each refresh
            time.sleep(2) 
            
        nameSignal = "Name Change Done"

        driver.close()
        driver.quit()
        del driver

        return nameSignal
        

def getOTP(userEmail:str, password:str) -> str:
    # Outlook IMAP settings
    outlook_server = "outlook.office365.com"

    # Connect to Outlook's IMAP server
    mail = imaplib.IMAP4_SSL(outlook_server)
    mail.login(userEmail, password)

    # Select the mailbox you want to access
    mail.select("inbox")

    # Search for all emails in the inbox
    status, messages = mail.search(None, "ALL")
    messages = messages[0].split()

    # Get the latest email
    latest_email_id = messages[-1]

    # Fetch the email by ID
    status, msg_data = mail.fetch(latest_email_id, "(RFC822)")
    raw_email = msg_data[0][1]

    # Parse the raw email
    msg = email.message_from_bytes(raw_email)

    # Extract the HTML part of the email
    for part in msg.walk():
        content_type = part.get_content_type()

        # Check for text/html parts
        if "text/html" in content_type:
            # Get the HTML content
            html_content = part.get_payload(decode=True).decode("utf-8")

    # Parse the HTML part with BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find the confirmation code element by its style attribute
    confirmation_code_element = soup.find('td', style='padding:10px;color:#565a5c;font-size:32px;font-weight:500;text-align:center;padding-bottom:25px;')

    # Extract the confirmation code
    if confirmation_code_element:
        confirmation_code = confirmation_code_element.get_text(strip=True)
        code = confirmation_code
    else:
        code = "Confirmation code element not found."
    
    # Close the connection
    mail.logout()

    #return the code
    return code


def writeOutputToFile(data: list, result_num: int) -> str:
    # Check if the file exists
    result_dir = './Scan_results'
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)

    file_path = os.path.join(result_dir, f'result{result_num}.xlsx')

    try:
        wb = load_workbook(file_path)
        ws = wb.active
    except FileNotFoundError:
        # If not, create a new workbook and add the headers
        wb = Workbook()
        ws = wb.active
        headers = ["Account", "Email change", "Name change"]
        for col_num, header in enumerate(headers,  1):
            col_letter = ws.cell(row=1, column=col_num).column_letter  # Get column letter
            ws['{}1'.format(col_letter)] = header

    # Append data starting from the second row
    ws.append(data)

    # Save the workbook
    wb.save(file_path)

    return file_path


def runBot(file_path: str = None):
    # Find the next available result file number
    result_num = 1
    while os.path.exists(f'./Scan_results/result{result_num}.xlsx'):
        result_num += 1

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook['Sheet1']  # Replace 'Sheet1' with the name of your sheet

    # Iterate through all rows in the sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        # 'values_only=True' returns the cell values instead of cell objects

        print(row)
        logging.info(f"INFO: {str(row)}", exc_info=True)
        
        CURRENT_USERNAME = row[0]
        CURRENT_PASSWORD = row[1]
        NEW_EMAIL_ADDRESS = row[2]
        NEW_EMAIL_PASSWORD = row[3]
        NEW_USERNAME = row[4]

        newBot = IGBot(CURRENT_USERNAME, CURRENT_PASSWORD, NEW_USERNAME, NEW_EMAIL_ADDRESS, NEW_EMAIL_PASSWORD)
        
        # Change email
        try:
            emailResponse = newBot.changeEmail()

        except Exception as e:
            
            if "net::ERR_NAME_NOT_RESOLVED" in str(e):
                emailResponse = "No internet or site not reachable. Try again"
            else:
                print(e)
                emailResponse = "Technical Difficulty, Check your list!"
            
        # Change name
        try:
            if emailResponse == "Instagram username or password incorrect":
                nameResponse = "Instagram username or password incorrect"
            else:
                nameResponse = newBot.changeName()

        except Exception as e:
            
            if "net::ERR_NAME_NOT_RESOLVED" in str(e):
                nameResponse = "No internet or site not reachable"
            else:
                print(e)
                nameResponse = "Technical Difficulty"
                
        # Create a list with the data
        data = [CURRENT_USERNAME, emailResponse, nameResponse]

        # Write the data to the file
        savedFile = writeOutputToFile(data, result_num)
    
    return savedFile


# For testing without the UI
if __name__ == "__main__":
    runBot("./datafile.txt")


            